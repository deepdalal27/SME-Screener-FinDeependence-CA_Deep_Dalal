#!/usr/bin/env python3
"""
Import YOUR OWN exports from Screener.in / Trendlyne / Tickertape / Excel.

Why this exists
---------------
Screener and Trendlyne have no public API and their terms prohibit scraping.
Scraping them would also be fragile and could get your users' IPs blocked.
But both let *you*, as a logged-in user, export data — and using your own
export is perfectly legitimate. This module reads those files and merges them
into the screener with full provenance, so a Screener-sourced number is
labelled "Screener.in export" in the app, not passed off as an exchange filing.

How to use
----------
1. On screener.in, build a screen that returns your SME universe
   (e.g. "Market Capitalization < 1000 AND Current Price > 0"), add whatever
   columns you want, then hit **Export to Excel**.
2. Drop the .xlsx / .csv into  data/import/
3. Run the normal build. Anything the exchanges couldn't give us gets filled
   from your export.

Any number of files can sit in that folder; later files win ties. Column names
are matched loosely, so Screener, Trendlyne and hand-made sheets all work
without editing headers. Unrecognised columns are ignored, never guessed at.
"""

import csv
import io
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import sources as S

CR = 1e7

# Map of normalised column header -> (our field, unit multiplier).
# Screener exports money in Rs Crore; ratios and percentages as plain numbers.
COLUMN_MAP = {
    # identity
    "name": ("name", None), "companyname": ("name", None), "company": ("name", None),
    "stockname": ("name", None),
    "nsecode": ("nse_symbol", None), "nsesymbol": ("nse_symbol", None),
    "symbol": ("nse_symbol", None), "ticker": ("nse_symbol", None),
    "bsecode": ("bse_code", None), "bsescripcode": ("bse_code", None),
    "isin": ("isin", None), "isincode": ("isin", None),
    "industry": ("industry", None), "sector": ("sector", None),
    # price / size
    "currentprice": ("price", 1), "price": ("price", 1), "cmp": ("price", 1),
    "lastprice": ("price", 1), "closeprice": ("price", 1),
    "marketcapitalization": ("mcap_cr", 1), "marketcap": ("mcap_cr", 1),
    "mcap": ("mcap_cr", 1), "marketcaprscr": ("mcap_cr", 1),
    # P&L (Rs Cr in the export -> rupees here)
    "sales": ("revenue", CR), "revenue": ("revenue", CR),
    "salesrscr": ("revenue", CR), "totalrevenue": ("revenue", CR),
    "revenuefromoperations": ("revenue", CR), "netsales": ("revenue", CR),
    "salespreviousyear": ("revenue_py", CR), "salesprecedingyear": ("revenue_py", CR),
    "operatingprofit": ("op_income", CR), "opm": ("_opm_pct", None),
    "opmpercent": ("_opm_pct", None), "operatingprofitmargin": ("_opm_pct", None),
    "ebitda": ("ebitda", CR), "ebit": ("ebit", CR),
    "depreciation": ("dep", CR), "interest": ("interest", CR),
    "financecost": ("interest", CR), "tax": ("tax", CR), "taxexpense": ("tax", CR),
    "profitaftertax": ("pat", CR), "netprofit": ("pat", CR), "pat": ("pat", CR),
    "netprofitrscr": ("pat", CR), "profitaftertaxpreviousyear": ("pat_py", CR),
    "eps": ("eps", 1), "epsrs": ("eps", 1), "earningspershare": ("eps", 1),
    # valuation
    "pricetoearning": ("_pe", None), "pe": ("_pe", None), "peratio": ("_pe", None),
    "pricetobook": ("_pb", None), "pb": ("_pb", None),
    "bookvalue": ("bvps", 1), "bookvaluepershare": ("bvps", 1),
    "dividendyield": ("div_yield", None),
    # cash flow
    "cashfromoperations": ("cfo", CR), "operatingcashflow": ("cfo", CR),
    "cashfromoperatingactivity": ("cfo", CR), "cfo": ("cfo", CR),
    "freecashflow": ("fcf", CR), "capitalexpenditure": ("capex", CR),
    "capex": ("capex", CR),
    # balance sheet / returns
    "reserves": ("_reserves", CR), "equitycapital": ("_equity_capital", CR),
    "totalequity": ("equity", CR), "shareholdersfunds": ("equity", CR),
    "borrowings": ("debt", CR), "totaldebt": ("debt", CR), "debt": ("debt", CR),
    "roce": ("_roce", None), "rocepercent": ("_roce", None),
    "returnoncapitalemployed": ("_roce", None),
    "roe": ("_roe", None), "returnonequity": ("_roe", None),
    "debttoequity": ("_de", None), "debtequity": ("_de", None),
    "promoterholding": ("promoter_pct", None), "promoterholdingpercent": ("promoter_pct", None),
    "pledgedpercentage": ("pledged_pct", None), "pledgedpercent": ("pledged_pct", None),
    # price data
    "highprice": ("hi52", 1), "high52week": ("hi52", 1), "week52high": ("hi52", 1),
    "lowprice": ("lo52", 1), "low52week": ("lo52", 1), "week52low": ("lo52", 1),
}

# Fields we accept as-is even though they are ratios the pipeline usually derives.
# They are only used when the pipeline could not derive them from raw figures.
DERIVED_HINTS = {"_pe", "_pb", "_roce", "_roe", "_de", "_opm_pct"}


def norm_col(s):
    """'Market Capitalization (Rs. Cr.)' -> 'marketcapitalization'"""
    s = re.sub(r"\(.*?\)", " ", str(s or ""))
    s = re.sub(r"(rs\.?\s*cr\.?|rs\.?|crore[s]?|percent|%)", " ", s, flags=re.I)
    return re.sub(r"[^a-z]", "", s.lower())


def read_table(path):
    """-> (headers, rows) from .csv/.tsv/.xlsx. Returns ([], []) if unreadable."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".csv", ".tsv", ".txt"):
            with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
                sample = f.read(4096)
                f.seek(0)
                delim = "\t" if (ext == ".tsv" or sample.count("\t") > sample.count(",")) else ","
                rows = list(csv.reader(f, delimiter=delim))
        elif ext in (".xlsx", ".xlsm"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                print(f"  ! {os.path.basename(path)}: needs openpyxl "
                      f"(pip install openpyxl) — skipped")
                return [], []
            wb = load_workbook(path, read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for r in ws.iter_rows(values_only=True):
                    rows.append(["" if v is None else v for v in r])
                if rows:
                    break            # first non-empty sheet
        else:
            return [], []
    except Exception as e:
        print(f"  ! could not read {os.path.basename(path)}: {e!r}")
        return [], []

    # find the header row: the first row where >=3 cells map to known columns
    for i, row in enumerate(rows[:25]):
        hits = sum(1 for c in row if norm_col(c) in COLUMN_MAP)
        if hits >= 3:
            return [norm_col(c) for c in row], rows[i + 1:]
    return [], []


def parse_file(path):
    """-> list of {field: value} dicts, one per company row."""
    headers, rows = read_table(path)
    if not headers:
        print(f"  ! {os.path.basename(path)}: no recognisable header row — skipped")
        return []
    out = []
    for row in rows:
        rec = {}
        for col, cell in zip(headers, row):
            spec = COLUMN_MAP.get(col)
            if not spec:
                continue
            field, mult = spec
            if mult is None and field in ("name", "nse_symbol", "bse_code", "isin",
                                          "industry", "sector"):
                txt = str(cell).strip()
                if txt and txt.lower() not in ("nan", "none", "-"):
                    rec[field] = txt
                continue
            v = S.num(cell)
            if v is None:
                continue
            rec[field] = v * mult if mult else v
        # a row is only useful if it identifies a company AND carries a figure
        if not any(rec.get(k) for k in ("nse_symbol", "bse_code", "isin", "name")):
            continue
        if not any(k for k in rec if k not in ("name", "nse_symbol", "bse_code",
                                               "isin", "industry", "sector")):
            continue
        out.append(rec)
    return out


def to_pipeline_values(rec):
    """Convert one imported row into the raw-figure dict the pipeline uses.
       Derived ratios are kept under their own keys so compute_metrics only
       falls back to them when it cannot calculate the figure itself."""
    vals, extra = {}, {}
    for k, v in rec.items():
        if k in ("name", "nse_symbol", "bse_code", "isin", "industry", "sector"):
            continue
        if k in DERIVED_HINTS:
            extra[k] = v
        else:
            vals[k] = v
    # equity = share capital + reserves, when the export gives those separately
    if "equity" not in vals and ("_reserves" in vals or "_equity_capital" in vals):
        eq = (vals.pop("_reserves", 0) or 0) + (vals.pop("_equity_capital", 0) or 0)
        if eq:
            vals["equity"] = eq
    vals.pop("_reserves", None)
    vals.pop("_equity_capital", None)
    # operating profit from OPM% x sales, only if OP wasn't given outright
    if "op_income" not in vals and extra.get("_opm_pct") is not None and vals.get("revenue"):
        vals["op_income"] = vals["revenue"] * extra["_opm_pct"] / 100.0
    return vals, extra


def load_imports(import_dir):
    """Read every file in data/import/ and index by each identifier present.
       -> {"NSE::SYM"|"BSE::code"|"ISIN"|"NAME::x": (values, extra, meta)}"""
    index = {}
    if not os.path.isdir(import_dir):
        return index
    files = sorted(f for f in os.listdir(import_dir)
                   if os.path.splitext(f)[1].lower() in (".csv", ".tsv", ".txt", ".xlsx", ".xlsm"))
    if not files:
        return index
    total = 0
    for fn in files:
        recs = parse_file(os.path.join(import_dir, fn))
        print(f"   {fn}: {len(recs)} rows")
        total += len(recs)
        for rec in recs:
            vals, extra = to_pipeline_values(rec)
            if not vals and not extra:
                continue
            meta = {k: rec[k] for k in ("name", "industry", "sector") if rec.get(k)}
            payload = (vals, extra, meta)
            if rec.get("isin"):
                index[rec["isin"].strip().upper()] = payload
            if rec.get("nse_symbol"):
                index["NSE::" + rec["nse_symbol"].strip().upper()] = payload
            if rec.get("bse_code"):
                code = re.sub(r"\D", "", str(rec["bse_code"]))
                if code:
                    index["BSE::" + code] = payload
            if rec.get("name"):
                index["NAME::" + S.name_key(rec["name"])] = payload
    print(f"   imported {total} rows from {len(files)} file(s)")
    return index


def match_import(company, index):
    """Find this company's imported row. Identifier matches first, name last."""
    if not index:
        return None
    isin = (company.get("isin") or "").upper()
    if isin and isin in index:
        return index[isin]
    sym = (company.get("symbol") or "").upper()
    if sym and company.get("exchange") == "NSE-EMERGE" and "NSE::" + sym in index:
        return index["NSE::" + sym]
    code = re.sub(r"\D", "", str(company.get("bse_code") or ""))
    if code and "BSE::" + code in index:
        return index["BSE::" + code]
    nm = S.name_key(company.get("name") or "")     # same key used when indexing
    if nm and "NAME::" + nm in index:
        return index["NAME::" + nm]
    return None


if __name__ == "__main__":
    d = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "import")
    print(f"Reading imports from {d}")
    idx = load_imports(d)
    print(f"indexed {len(idx)} lookup keys")
